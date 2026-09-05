#!/usr/bin/env python3

from __future__ import annotations

import html
import re
import tempfile
import time
from datetime import date, datetime
from pathlib import Path

import requests
from openpyxl import load_workbook

SPREADSHEET_ID = "1ivz77GAXHLpLbAODANGr8l09cq2Bo6TVfcoYUS3o7lM"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"

HERE = Path(__file__).resolve().parent
OUTPUT = HERE / "index.html"

TEACHER_SHEET = "Teacher Calendar"
PACING_SHEET = "Pacing"
DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
MARKER_COLS = [2, 4, 6, 8, 10]   # B, D, F, H, J
CONTENT_COLS = [3, 5, 7, 9, 11]  # C, E, G, I, K

# Keep this list in sync with the Phy Resources sheet.
RESOURCE_LINKS = [
    ('Overview', 'https://docs.google.com/document/d/1rrToxZ84-VGe-75JeIofcH6FXqdCFiE-MNlwDZNvMs4/edit?usp=sharing'),
    ('Website', 'https://tnezki.github.io/physics/'),
    ('Agenda', 'https://tnezki.github.io/physics/agenda/index.html'),
    ('Textbook', 'https://tnezki.github.io/textbooks/phy/index.html'),
    ('Formula Sheet', 'https://tnezki.github.io/physics/misc/formula_sheet.html'),
    ('Printables', 'https://tnezki.github.io/physics/misc/printables/aaagallery_index.html'),
    ('Desmos', 'https://www.desmos.com/calculator'),
    ('Canvas', 'https://mariners.instructure.com/'),
    ('Upload Spot', 'https://drive.google.com/drive/folders/1wxxAxIxJ9yU5goiVNriIVpmiEJmE7SX2?usp=drive_link'),
    ('PhET', 'https://phet.colorado.edu/en/simulations/filter?subjects=physics'),
    ('oPhysics', 'https://ophysics.com/index.html'),
    ('Walter F', 'https://www.walter-fendt.de/html5/phen/'),
    ('Falsted', 'https://www.falstad.com//mathphysics.html'),
    ('Lewin Videos', 'https://www.youtube.com/channel/UCiEHVhv0SBMpP75JbzJShqw'),
    ('Hewitt Videos', 'https://conceptual.academy/'),
]


def safe_text(value):
    if value is None or value.__class__.__name__ == "ArrayFormula":
        return ""
    if isinstance(value, (datetime, date)):
        return f"{value.month}/{value.day}"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalized_key(value):
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def normalized_label(value):
    return re.sub(r"\s+", " ", safe_text(value)).strip().casefold()


def is_x(value):
    return safe_text(value).casefold() == "x"


def looks_like_date(value):
    if isinstance(value, (datetime, date)):
        return True
    if value is None:
        return False
    return bool(re.fullmatch(r"\d{1,2}/\d{1,2}(?:/\d{2,4})?", safe_text(value)))


def date_key(value, school_start_year=2026):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None

    match = re.fullmatch(
        r"(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?",
        safe_text(value),
    )
    if not match:
        return None

    month = int(match.group(1))
    day = int(match.group(2))
    year_text = match.group(3)

    if year_text:
        year = int(year_text)
        if year < 100:
            year += 2000
    else:
        year = school_start_year if month >= 7 else school_start_year + 1

    try:
        return date(year, month, day)
    except ValueError:
        return None


def direct_cell_link(cell):
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target

    formula = cell.value
    if isinstance(formula, str) and formula.startswith("="):
        match = re.search(r'HYPERLINK\(\s*"([^"]+)"', formula, re.I)
        if match:
            return match.group(1)

    return None


def download_workbook():
    # A unique query string avoids a stale Google export being reused by an
    # intermediary cache. Cache-control headers provide a second safeguard.
    url = f"{EXPORT_URL}&cachebust={time.time_ns()}"
    response = requests.get(
        url,
        timeout=45,
        headers={
            "User-Agent": "Mozilla/5.0 AgendaBuilder/2.1",
            "Cache-Control": "no-cache, no-store, max-age=0",
            "Pragma": "no-cache",
        },
    )
    response.raise_for_status()
    if len(response.content) < 1000:
        raise RuntimeError("Google returned an unexpectedly small workbook export.")

    handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    handle.write(response.content)
    handle.close()
    return Path(handle.name)


def build_pacing_link_lookup(pacing_values, pacing_formulas):
    lookup = {}
    direct_link_count = 0

    for row in range(1, min(pacing_values.max_row, 500) + 1):
        pacing_key = normalized_key(pacing_values.cell(row=row, column=1).value)
        if pacing_key in (None, ""):
            continue

        row_links = lookup.setdefault(pacing_key, {})

        for col in range(2, min(pacing_values.max_column, 26) + 1):
            label = safe_text(pacing_values.cell(row=row, column=col).value)
            if not label:
                continue

            link = (
                direct_cell_link(pacing_formulas.cell(row=row, column=col))
                or direct_cell_link(pacing_values.cell(row=row, column=col))
            )
            if link:
                direct_link_count += 1
                row_links.setdefault(normalized_label(label), link)

    return lookup, direct_link_count


def find_teacher_weeks(teacher_values):
    """Find weekly blocks from the five date cells in C/E/G/I/K."""
    date_rows = []

    for row in range(1, teacher_values.max_row + 1):
        values = [teacher_values.cell(row=row, column=col).value for col in CONTENT_COLS]
        if sum(1 for value in values if looks_like_date(value)) >= 4:
            date_rows.append(row)

    if not date_rows:
        raise RuntimeError("Could not find weekly date rows in Teacher Calendar.")

    weeks = []
    for index, date_row in enumerate(date_rows):
        pacing_row = date_row - 1
        next_date_row = date_rows[index + 1] if index + 1 < len(date_rows) else None
        content_start = date_row + 1
        content_end = (
            next_date_row - 2
            if next_date_row is not None
            else min(date_row + 25, teacher_values.max_row)
        )

        dates_raw = [teacher_values.cell(row=date_row, column=col).value for col in CONTENT_COLS]
        dates = [safe_text(value) for value in dates_raw]
        date_keys = [date_key(value) for value in dates_raw]
        pacing_keys = [
            normalized_key(teacher_values.cell(row=pacing_row, column=col).value)
            for col in CONTENT_COLS
        ]

        week_number = safe_text(teacher_values.cell(row=date_row, column=1).value)
        manual_current = any(
            is_x(teacher_values.cell(row=row, column=1).value)
            for row in range(pacing_row, content_end + 1)
        )

        weeks.append(
            {
                "week_number": week_number,
                "pacing_row": pacing_row,
                "date_row": date_row,
                "content_start": content_start,
                "content_end": content_end,
                "dates_raw": dates_raw,
                "dates": dates,
                "date_keys": date_keys,
                "pacing_keys": pacing_keys,
                "manual_current": manual_current,
            }
        )

    return weeks


def choose_current_week(weeks, today=None):
    today = today or date.today()

    marked = [index for index, week in enumerate(weeks) if week["manual_current"]]
    if len(marked) == 1:
        return marked[0]
    if len(marked) > 1:
        print("Warning: more than one Teacher Calendar week has x in column A; using date instead.")

    for index, week in enumerate(weeks):
        valid = [d for d in week["date_keys"] if d is not None]
        if valid and min(valid) <= today <= max(valid):
            return index

    earlier = []
    for index, week in enumerate(weeks):
        valid = [d for d in week["date_keys"] if d is not None]
        if valid and min(valid) <= today:
            earlier.append((min(valid), index))

    if earlier:
        return max(earlier)[1]

    return 0


def resolve_item_link(teacher_value_cell, teacher_formula_cell, label, pacing_key, pacing_links):
    direct = direct_cell_link(teacher_formula_cell) or direct_cell_link(teacher_value_cell)
    if direct:
        return direct

    if pacing_key not in (None, "") and label:
        return pacing_links.get(pacing_key, {}).get(normalized_label(label))

    return None


def read_week_rows(week, teacher_values, teacher_formulas, pacing_links):
    """Publish checked nonblank items and compact each day independently."""
    day_items = [[] for _ in CONTENT_COLS]

    for row in range(week["content_start"], week["content_end"] + 1):
        for day_index, (marker_col, content_col) in enumerate(zip(MARKER_COLS, CONTENT_COLS)):
            selected = is_x(teacher_values.cell(row=row, column=marker_col).value)
            label = safe_text(teacher_values.cell(row=row, column=content_col).value)

            if not selected or not label:
                continue

            url = resolve_item_link(
                teacher_values.cell(row=row, column=content_col),
                teacher_formulas.cell(row=row, column=content_col),
                label,
                week["pacing_keys"][day_index],
                pacing_links,
            )
            day_items[day_index].append((label, url))

    row_count = max((len(items) for items in day_items), default=0)
    rows = [
        [items[row_index] if row_index < len(items) else None for items in day_items]
        for row_index in range(row_count)
    ]
    return rows, day_items


def read_calendar_from_path(xlsx, today=None):
    wb_values = load_workbook(xlsx, data_only=True, read_only=False)
    wb_formulas = load_workbook(xlsx, data_only=False, read_only=False)

    for sheet_name in (TEACHER_SHEET, PACING_SHEET):
        if sheet_name not in wb_values.sheetnames:
            raise RuntimeError(f"Missing sheet: {sheet_name}")

    teacher_values = wb_values[TEACHER_SHEET]
    teacher_formulas = wb_formulas[TEACHER_SHEET]
    pacing_values = wb_values[PACING_SHEET]
    pacing_formulas = wb_formulas[PACING_SHEET]

    pacing_links, direct_link_count = build_pacing_link_lookup(
        pacing_values,
        pacing_formulas,
    )
    weeks = find_teacher_weeks(teacher_values)
    current_index = choose_current_week(weeks, today=today)

    for week in weeks:
        week["rows"], week["day_items"] = read_week_rows(
            week,
            teacher_values,
            teacher_formulas,
            pacing_links,
        )

    current = weeks[current_index]

    print(f"Teacher Calendar weekly blocks found: {len(weeks)}")
    print(
        f"Current week: {current.get('week_number') or '?'} | "
        + ", ".join(current["dates"])
    )
    print(f"Pacing direct hyperlinks available: {direct_link_count}")
    print(f"All weeks listed below current week: {len(weeks)}")
    for day_name, items in zip(DAY_NAMES, current["day_items"]):
        labels = ", ".join(label for label, _ in items) or "(none)"
        print(f"  {day_name}: {labels}")

    return {
        "current": current,
        "current_index": current_index,
        "all_weeks": weeks,
    }


def read_calendar():
    xlsx = download_workbook()
    try:
        return read_calendar_from_path(xlsx)
    finally:
        try:
            xlsx.unlink()
        except OSError:
            pass


def render_link(label, url, kind=""):
    classes = "cal-link" + (f" {kind}" if kind else "")
    if url:
        return (
            f'<a class="{classes}" href="{html.escape(url, quote=True)}" '
            f'target="_blank" rel="noopener">{html.escape(label)}</a>'
        )
    return f'<span class="{classes} no-link">{html.escape(label)}</span>'


def item_kind(label, row_index):
    if row_index != 0:
        return ""
    low = label.casefold()
    if any(term in low for term in ("labor day", "pd", "no school", "holiday", "break")):
        return "holiday"
    return "lesson"


def render_week(week, current=False, list_state=None):
    if current:
        cells = "".join(
            f"<th><span class='dow'>{day}</span><span class='date'>{html.escape(day_date)}</span></th>"
            for day, day_date in zip(DAY_NAMES, week["dates"])
        )
        cls = "current-week"
    else:
        cells = "".join(
            f"<th><div class='date'>{html.escape(day_date)}</div></th>"
            for day_date in week["dates"]
        )
        cls = {
            "past": "previous-week",
            "current": "all-current-week",
            "future": "previous-week",
        }.get(list_state, "previous-week")

    header = f'<tr class="week-head">{cells}</tr>'
    body_rows = []

    for row_index, row_cells in enumerate(week["rows"]):
        tds = []
        for item in row_cells:
            if item is None:
                tds.append("<td></td>")
                continue

            label, url = item
            kind = item_kind(label, row_index)
            tds.append(f"<td>{render_link(label, url, kind)}</td>")

        body_rows.append("<tr>" + "".join(tds) + "</tr>")

    if not body_rows:
        body_rows.append('<tr><td colspan="5" class="empty-week">No published agenda items.</td></tr>')

    return f'<tbody class="week-block {cls}">{header}{"".join(body_rows)}</tbody>'


def build_html(calendar):
    resources = "\n".join(
        f'<a href="{html.escape(url, quote=True)}" target="_blank" rel="noopener">{html.escape(label)}</a>'
        for label, url in RESOURCE_LINKS
    )

    current_html = render_week(calendar["current"], current=True)

    all_weeks_html = (
        '<tbody class="previous-weeks-divider"><tr>'
        '<td colspan="5">ALL WEEKS</td>'
        '</tr></tbody>'
        + "".join(
            render_week(
                week,
                current=False,
                list_state=(
                    "past" if index < calendar["current_index"]
                    else "current" if index == calendar["current_index"]
                    else "future"
                ),
            )
            for index, week in enumerate(calendar["all_weeks"])
        )
    )

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta http-equiv="refresh" content="300">
<title>Physics Agenda 2026-2027</title>
<style>
:root {{
  --navy:#173f6d;
  --navy-dark:#173f6d;
  --gold:#e0bd4f;
  --gold-light:#fff0b8;
  --gold-pale:#fff8df;
  --current-row-a:#fffaf0;
  --current-row-b:#fff1bd;
  --ink:#1f2937;
  --muted:#64748b;
  --lesson:#fff0b8;
  --link:#173f6d;
}}
* {{ box-sizing:border-box; }}
body {{
  margin:0;
  font-family:Arial, Helvetica, sans-serif;
  background:#fff;
  color:var(--ink);
}}
.wrapper {{
  width:min(980px, calc(100% - 24px));
  margin:18px auto 40px;
}}
.titlebar {{
  background:var(--navy);
  color:#fff;
  padding:7px 11px;
  border-radius:10px 10px 0 0;
  text-align:center;
}}
.titlebar h1 {{
  margin:0;
  font-size:.88rem;
  font-weight:800;
}}
.titlebar .small {{
  font-size:1em;
  font-weight:600;
}}
.resources {{
  border:1px solid var(--gold);
  border-top:0;
  padding:12px 14px 13px;
  display:flex;
  flex-wrap:wrap;
  justify-content:center;
  align-items:center;
  text-align:center;
  gap:8px;
  background:#fff9e9;
}}
.resources a {{
  text-decoration:none;
  color:var(--navy-dark);
  background:#fff;
  border:1px solid var(--gold);
  border-radius:999px;
  padding:7px 11px;
  font-size:.88rem;
  font-weight:700;
}}
.resources a:hover,
.resources a:focus-visible {{
  background:var(--gold-light);
  border-color:var(--navy);
}}
.calendar-wrap {{
  overflow-x:auto;
  border:1px solid #c8b675;
  border-top:0;
}}
table {{
  border-collapse:collapse;
  width:100%;
  max-width:100%;
  min-width:0;
  table-layout:fixed;
}}
th, td {{
  border-right:1px solid #cfd4da;
  border-bottom:1px solid #cfd4da;
  text-align:center;
  vertical-align:middle;
}}
tr > *:last-child {{ border-right:0; }}
.week-head th {{ width:20%; padding:7px 5px; }}
.dow {{ font-size:.82rem; font-weight:800; }}
.date {{ margin-top:2px; font-size:.76rem; font-weight:700; }}
td {{ padding:4px 5px; background:#fff; height:32px; }}
.cal-link {{
  display:block;
  width:100%;
  text-decoration:none;
  color:var(--link);
  font-size:.82rem;
  font-weight:650;
  padding:4px 4px;
  border-radius:5px;
  overflow-wrap:anywhere;
}}
a.cal-link:hover,
a.cal-link:focus-visible {{
  background:#fff4c7;
  text-decoration:none;
}}
.cal-link.lesson {{
  background:var(--lesson);
  border:1px solid #e1c86d;
  font-weight:800;
  color:var(--navy-dark);
}}
.cal-link.holiday {{
  background:#f6edcf;
  color:#475569;
  font-weight:800;
}}
.no-link {{ cursor:default; }}
.empty-week {{ color:var(--muted); font-size:.8rem; padding:10px; }}

/* Featured current week: compact navy day/date header + light gold alternating rows. */
.current-week .week-head th {{
  background:var(--navy);
  color:#fff;
  text-align:left;
  padding:7px 6px;
}}
.current-week .dow {{
  display:inline;
  font-size:.86rem;
  font-weight:850;
  line-height:1.08;
}}
.current-week .date {{
  display:inline;
  margin:0 0 0 6px;
  font-size:.86rem;
  font-weight:900;
  color:#fff;
}}
.current-week td {{
  padding:0;
  height:40px;
  min-height:40px;
  background:var(--current-row-a);
}}
.current-week tr:nth-child(odd):not(.week-head) td {{
  background:var(--current-row-b);
}}
.current-week .cal-link {{
  display:flex;
  align-items:center;
  justify-content:center;
  width:100%;
  min-height:40px;
  padding:7px 4px;
  font-size:.90rem;
  font-weight:750;
  line-height:1.15;
}}
.current-week .cal-link.lesson {{
  background:rgba(255,255,255,.45);
  border:1px solid #dbc36d;
  color:var(--navy);
  font-size:.96rem;
  font-weight:900;
  text-decoration:none;
}}
.current-week .cal-link.holiday {{
  font-size:.90rem;
  font-weight:900;
}}

.previous-weeks-divider td {{
  background:var(--navy) !important;
  color:#fff;
  font-size:.96rem;
  font-weight:800;
  padding:9px 7px;
}}

/* The current week is highlighted again inside ALL WEEKS. */
.all-current-week .week-head th {{
  background:var(--gold);
  color:var(--navy);
  border-top:5px solid var(--navy);
  text-align:left;
  padding-left:9px;
}}
.all-current-week .week-head th .date {{
  color:var(--navy);
  font-size:.9rem;
  font-weight:900;
}}
.all-current-week tr td {{
  background:var(--current-row-a) !important;
  border-color:#d8c77f;
}}
.all-current-week tr:nth-child(even) td {{
  background:var(--current-row-b) !important;
}}
.all-current-week .cal-link.lesson {{
  background:#fff6d5;
  border:1px solid #dbc36d;
  color:var(--navy);
  font-weight:900;
}}
.all-current-week .cal-link.holiday {{
  background:#f6edcf;
  color:#4b5563;
}}
.all-current-week a.cal-link:hover,
.all-current-week a.cal-link:focus-visible {{
  background:#ffe99b;
}}

/* Every non-current week stays in the compact gray reference style. */
.previous-week .week-head th {{
  background:#3f4650;
  color:#fff;
  border-top:5px solid #20242a;
  text-align:left;
  padding-left:9px;
}}
.previous-week .week-head th .date {{ color:#e5e7eb; }}
.previous-week tr td {{
  background:#fff !important;
  border-color:#cfd4da;
}}
.previous-week tr:nth-child(even) td {{
  background:#f3f4f6 !important;
}}
.previous-week .cal-link.lesson {{
  background:#e5e7eb;
  border:1px solid #c7ccd1;
  color:#2f3740;
}}
.previous-week .cal-link.holiday {{
  background:#eceff2;
  color:#4b5563;
}}
.previous-week a.cal-link:hover,
.previous-week a.cal-link:focus-visible {{
  background:#e2e6ea;
}}
.updated {{
  text-align:center;
  color:var(--muted);
  font-size:.76rem;
  padding-top:8px;
}}

@media (max-width:700px) {{
  .wrapper {{ width:100%; margin:0; }}
  .titlebar {{ border-radius:0; }}
  .resources {{ justify-content:center; padding:10px; }}
  .resources a {{ font-size:.8rem; padding:6px 9px; }}
}}
</style>
</head>
<body>
<div class="wrapper">
  <header class="titlebar">
    <h1>Physics <span class="small">– Agenda 2026-2027</span></h1>
  </header>

  <nav class="resources" aria-label="Student resources">
    {resources}
  </nav>

  <div class="calendar-wrap">
    <table aria-label="Physics student agenda">
      {current_html}
      {all_weeks_html}
    </table>
  </div>

  <div class="updated">Agenda generated {html.escape(stamp)}</div>
</div>
</body>
</html>
"""


def main():
    calendar = read_calendar()
    output = build_html(calendar)
    temp = OUTPUT.with_suffix(".html.tmp")
    temp.write_text(output, encoding="utf-8")
    temp.replace(OUTPUT)
    print(f"Updated {OUTPUT}")


if __name__ == "__main__":
    main()
